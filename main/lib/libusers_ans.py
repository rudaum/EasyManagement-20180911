#!/usr/bin/python
"""
- Purpose:
	Using Ansible as backbone, this library provides functions to
	retrieve a list of users of one or more valid hosts registered
	in the inventory of ansible. This list of users can then be
	parsed to generated excel sheets or as source to Flask

- Author:
	Rudolf Wolter (KN OSY Team)

- Contact for questions and/or comments:
	rudolf.wolter@kuehne-nagel.com

- Version Releases and modifications.
	> 1.0 (30.08.2017) - Initial release with core functionalities.
	> 1.0.1 (12.09.2017) - Implemented 'format_wb' improved 'mksheet'

- TODO:
	Retrieve and place somewhere the last login
	DONE - lsusers not working properly when uising 'user_filter'
"""
### START OF MODULE IMPORTS 
# --------------------------------------------------------------- #
from subprocess import Popen, PIPE
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from collections import OrderedDict, MutableSet
from pandas import read_excel
# --------------------------------------------------------------- #
### END OF MODULE IMPORTS

### START OF GLOBAL VARIABLES DECLARATION
# --------------------------------------------------------------- #
CONFLICT_COUNTER = 0
WHITE_COLOR = 'FFFFFF'
HEADER_COLOR = '7EA9DE'
SPOT_COLOR = 'D6A6A6'

HEADER_FILL = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type='solid')
ATTR_FILL = PatternFill(start_color=WHITE_COLOR, end_color=WHITE_COLOR, fill_type='solid')
SPOT_FILL = PatternFill(start_color=SPOT_COLOR, end_color=SPOT_COLOR, fill_type='solid')
HEADER_FONT = Font(name='Calibri', size=11, bold=True, color=WHITE_COLOR)
HYPERLINK_FONT = Font(name='Calibri', size=11, underline='single', color='FF0000')
HEADER_BORDER = Border(left=Side(style='thin'),
					   right=Side(style='thin'),
					   top=Side(style='thin'),
					   bottom=Side(style='thin'))

FILENAME = 'em_lsusers.xlsx'
# --------------------------------------------------------------- #
### END OF GLOBAL VARIABLES DECLARATION

### START OF CLASS DEFINITIONS
# --------------------------------------------------------------- #
class UserHtmlTable():
    def __init__(self, userdict):
        self.l = 0
        self.userdict = userdict
        self.table_code = []

        self.append_code('<table class="table table-hover table-responsive">')
        #- Table Header code
        self.append_code('<thead>')
        self.append_code('<tr>')
        for line_code in self.get_serversCode(): # Inserting the Server names as Table Header
            self.append_code(line_code)
        self.append_code('</tr>')
        self.append_code('</thead>')
        #- End of Table Header Code

        #- Table Body code and attribute rows
        self.append_code('<tbody>')
        for attr in self.get_userAttr():  # Inserting the Attributes and values
            for line_code in self.get_attrCode(attr):
                self.append_code(line_code)
        self.append_code('</tbody>')
        # - End of Table Body Code
        self.append_code('</table>')

    def append_code(self, code):
        if code[0] == '<' and code[1] != '/' and '</' in code:
            self.table_code.append(t(self.l) + code)

        elif code[0] == '<' and code[1] == '/':
            self.l -= 1
            self.table_code.append(t(self.l) + code)

        elif code[0] == '<' and code[1] != '/':
            self.table_code.append(t(self.l) + code)
            self.l += 1
        else:
            print 'invalid HTML code'

    def get_serversCode(self):
        srvcode = []
        srvcode.append('<th></th>') # Header's first column hast to be empty
        for server in self.userdict.keys():
            srvcode.append('<th>{}</th>'.format(server))
        return srvcode

    def get_userAttr(self):
        attrs = list()
        for srv in self.userdict.keys():
            for attr in self.userdict[srv]:
                attrs.append(attr)
        attrs = list(OrderedDict.fromkeys(attrs))
        return attrs

    def get_attrCode(self,attr):
        code = []
        for server in self.userdict.keys():
            code.append('<td>{}</td>'.format(self.userdict[server][attr]))
        if len(set(code)) > 1:
            code.insert(0,'<tr class="danger">')
        else:
            code.insert(0, '<tr>')

        code.insert(1,'<th>{}</th>'.format(attr))
        code.append('</tr>')
        return code

    def get_code(self):
        for code_line in self.table_code:
            print code_line


# --------------------------------------------------------------- #
### END OF CLASS DEFINITIONS

### START OF FUNCTIONS DECLARATION
# --------------------------------------------------------------- #
def find_in_col(ws, col, search_str):
    for col in ws.iter_cols(min_col=col, max_col=col, max_row=ws.max_row):
        for cell in col:
            if cell.value == search_str:
                return cell
    return False
# --------------------------------------------------------------- #
# --------------------------------------------------------------- #
def find_in_row(ws, row, search_str):
    for row in ws.iter_rows(min_row=row, max_col=ws.max_column, max_row=row):
        for cell in row:
            if cell.value == search_str:
                return cell
    return False
# --------------------------------------------------------------- #
# --------------------------------------------------------------- #
def get_max_col(ws, row):
    max_col = 1
    for col in ws.iter_cols(min_col=2, max_col=8000, min_row=row, max_row=row):
        for cell in col:
            if cell.value is None:
                return max_col
            else:
                max_col += 1
# --------------------------------------------------------------- #
# --------------------------------------------------------------- #
def order_by_user(raw_users):
    users = OrderedDict()
    for host in raw_users.iterkeys():  # Hostname Level
        for user, attr in raw_users[host]["users"].iteritems():  # User level
            if user not in users:
                users[user] = OrderedDict()
            if host not in users[user]:
                users[user][host] = OrderedDict()
            users[user][host] = attr
    return users
# --------------------------------------------------------------- #

def t(n):
    '''
    To return a number of tabulations, given the paramenter n
    :param n - tab multipler. Use 0 to none
    '''
    return ''.join(['    '] * n)
# --------------------------------------------------------------- #
def mk_html_table(user_data):
    l = 0
    table_code = []
    opened_tags = []
    closed_tags = []

    def append_code(code):
        l=0
        if code[0] == '<' and code[1] == '/':
            closed_tags.append(code.split(' ')[0].strip('</'))
            table_code.append(t(l)+code)
            l = - 1
        elif code[0] == '<' and code[1] != '/':
            opened_tags.append(code.split(' ')[0].strip('</'))
            table_code.append(t(l)+code)
            l = + 1
        else:
            print 'invalid HTML code'

    append_code('<table class="table table-hover table-responsive">')
    append_code('<thead>')
    append_code('<tr>')

    return table_code
# --------------------------------------------------------------- #


# --------------------------------------------------------------- #
def update_report(wb, user, attr, attr_ref):
	global CONFLICT_COUNTER
	ws = wb.worksheets[0]
	if find_in_col(ws, 1, user) is False:
		targ_row = ws.max_row + 2
		ws.cell(row=targ_row, column=1, value=user)
		ws.cell(row=targ_row, column=1).fill = HEADER_FILL
		ws.cell(row=targ_row, column=1).font = HEADER_FONT

	targ_row = ws.max_row + 1
	CONFLICT_COUNTER += 1
	link_ref = '{}!{}'.format(user,attr_ref)
	conflict = '=HYPERLINK("#{}","Potential conflict spotted for: {}")'.format(link_ref,attr)
	ws.cell(row=targ_row, column=2, value=conflict)
	ws.cell(row=targ_row, column=2).font = HYPERLINK_FONT
	ws.cell(row=targ_row, column=2).fill = SPOT_FILL
# --------------------------------------------------------------- #

# --------------------------------------------------------------- #
def format_wb(wb):
	for sheet_index in range(1, len(wb.sheetnames)):
		ws = wb.worksheets[sheet_index] #getting user's sheet
		ws.column_dimensions['A'].width = 17
		spotted_rows = set()
		for row in range(1, ws.max_row + 1):
			#ws_max_col = get_max_col(ws, row)
			for col in ws.iter_cols(min_row=row, max_row=row, min_col=1, max_col=ws.max_column):
				for cell in col:
					if cell.row is 1 or cell.column is 'A':
						cell.fill = HEADER_FILL
						cell.border = HEADER_BORDER
						cell.font = HEADER_FONT
						if cell.column is not 'A':
							ws.column_dimensions[cell.column].width = 19
					elif cell.row > 1 and cell.column is not 'A':
						cell.fill = ATTR_FILL
						if cell.column is not 'B' :
							# Checking for possible conflicts:
							if cell.value != ws['B' + str(cell.row)].value:
								spotted_rows.add(cell.row) # Adding possible conflict

		# Handling conflicts found
		for row in spotted_rows:
			attr_ref = 'A'+str(row) #setting the attribute's name reference cell
			update_report(wb, ws.title, ws[attr_ref].value, attr_ref) #Updating 'Report' sheet
			for col in ws.iter_cols(min_row=row, max_row=row, min_col=1, max_col=ws.max_column):
				for cell in col:
					cell.fill = SPOT_FILL

	ws = wb.worksheets[0]
	ws.title = 'Report'
	ws['A1'] = '{} Potential Inconsistencies Found:'.format(CONFLICT_COUNTER)
	ws.column_dimensions['A'].width = 15
	ws.column_dimensions['B'].width = 42
	ws.merge_cells('A1:B1')
	ws['A1'].fill = HEADER_FILL
	ws['A1'].border = HEADER_BORDER
	ws['A1'].font = Font(name='Calibri', size=14, bold=True, color=WHITE_COLOR)
	ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
# --------------------------------------------------------------- #
# --------------------------------------------------------------- #

def lsusers(targ_hosts, fulllist=False, user_filter="ALL"):
	"""
	Returns user(s) atributes from the target hosts.
	It contructs an ansible command that is called using subprocess
	module. Then it retrieves a user list with their attributes
	from all target hosts and formats it as Dictionary tree
	"""
	hosts = OrderedDict()

	if fulllist:
		ans_cmd = ["/bin/ansible", "-ba", "lsuser -f " + user_filter]
	else:
		ans_cmd = ["/bin/ansible", "-a", "lsuser -f " + user_filter]
	ans_cmd.append(targ_hosts)

	# Calling Ansible process
	output = Popen(ans_cmd, stdout=PIPE, stderr=PIPE)
	msg_handler = ''

	# Parsing the process output
	for line in output.stdout:
		if " | " in line and "rc=" in line and ">>" in line:  # Parsing hosts line
			hn = line.split(" | ")[0]
			hosts[hn] = OrderedDict()
			hosts[hn]["exec_rc"] = line.split(" | ")[2].split(" ")[0]
			hosts[hn]["exec_msg"] = line.split(" | ")[1]
			hosts[hn]["users"] = OrderedDict()

		# Checking if Ansible could reach/access the host
		# if not, set msg_handler to get the error message in the net line and delete the host's key
		elif '| SUCCESS |' not in line and " | " in line and " => {" in line:  # Parsing Host / Target
			msg_handler = 'HOSTFAIL'
			hn = line.split(" | ")[0]
			hosts[hn] = OrderedDict()
			hosts[hn]["exec_rc"] = line.split(" | ")[1].split(" ")[0]
			hosts[hn]["exec_msg"] = line.split(" | ")[1].split(" ")[0]
			hosts[hn]["users"] = OrderedDict()

		elif ":\n" in line and msg_handler == '':  # Parsing Users Line
			user = line.strip()[:-1]  # stripping ":" from the user
			hosts[hn]["users"][user] = OrderedDict()

		elif "=" in line and msg_handler == '':  # Parsing Attribute
			an = line.strip().split("=", 1)[0]  # attribute's name
			if '_last_' not in an and 'unsuccessful_' not in an:  # Filtering Undesired  attribures
				av = line.strip().split("=", 1)[1].decode('utf-8')  # attribute's value
				hosts[hn]["users"][user][an] = av
		else:
			# Could not reach the server, display the error message and deletes the Key from the dictionary
			if msg_handler == 'HOSTFAIL':
				print 'WARNING! Host {} failed with message: {}'.format(hn, hosts[hn]['exec_msg'])
				del hosts[hn]['exec_rc']
				del hosts[hn]['exec_msg']
				del hosts[hn]['users']
				del hosts[hn]
				msg_handler = ''

	return hosts
# --------------------------------------------------------------- #
# --------------------------------------------------------------- #
def mksheet(raw_users):
	"""
	Builds an excel sheet out from the 'lsusers' function output,
	which comes as a dictionary tree. Then it saves it as a xlsx
	file format
	"""
	users = order_by_user(raw_users)  # formating lsusers raw output
	wb = Workbook()  # creating new workbook
	for user in users.iterkeys():  # Looping over Users
		ws = wb.create_sheet(user)  # Creating user's Tab
		for host in users[user].iterkeys():  # Looping over hosts
			host_col = ws.max_column + 1
			ws.cell(row=1, column=host_col, value=host)  # hostname on 1st row
			ws.cell(row=2, column=1, value="User")  # User name Attribute
			ws.cell(row=2, column=host_col, value=user)  # User name Value
			for attr in users[user][host].iteritems():  # Looping over attributes
				find_attr = find_in_col(ws, 1, attr[0])
				if find_attr is False:
					attr_row = ws.max_row + 1
					ws.cell(row=attr_row, column=1, value=attr[0])  # writting attribute Name
				else:
					attr_row = find_attr.row

				ws.cell(row=attr_row, column=host_col, value=attr[1])  # writting attribute value

	format_wb(wb)
	wb.save(FILENAME)  # saving the workbook
	wb.close()
# --------------------------------------------------------------- #
# --------------------------------------------------------------- #
def mkhtml():
	"""
	Builds an html Jinja2 Template out from the 'lsusers' function output,
	which comes as a dictionary tree.
	"""

	raw_users = OrderedDict([('dehamsp40', OrderedDict([('exec_rc', 'rc=0'), ('exec_msg', 'SUCCESS'), ('users', OrderedDict([('root', OrderedDict([('id', u'0'), ('pgrp', u'system'), ('groups', u'system,bin,sys,security,cron,audit,lp'), ('home', u'/'), ('shell', u'/usr/bin/ksh')])), ('daemon', OrderedDict([('id', u'1'), ('pgrp', u'staff'), ('groups', u'staff'), ('home', u'/etc')])), ('bin', OrderedDict([('id', u'2'), ('pgrp', u'bin'), ('groups', u'bin,sys,adm'), ('home', u'/bin')])), ('sys', OrderedDict([('id', u'3'), ('pgrp', u'sys'), ('groups', u'sys'), ('home', u'/usr/sys')])), ('adm', OrderedDict([('id', u'4'), ('pgrp', u'adm'), ('groups', u'adm'), ('home', u'/var/adm')])), ('uucp', OrderedDict([('id', u'5'), ('pgrp', u'uucp'), ('groups', u'uucp'), ('home', u'/usr/lib/uucp')])), ('guest', OrderedDict([('id', u'100'), ('pgrp', u'usr'), ('groups', u'usr'), ('home', u'/home/guest')])), ('nobody', OrderedDict([('id', u'4294967294'), ('pgrp', u'nobody'), ('groups', u'nobody'), ('home', u'/')])), ('lpd', OrderedDict([('id', u'9'), ('pgrp', u'nobody'), ('groups', u'nobody'), ('home', u'/')])), ('lp', OrderedDict([('id', u'11'), ('pgrp', u'lp'), ('groups', u'lp,printq'), ('home', u'/var/spool/lp'), ('shell', u'/bin/false')])), ('invscout', OrderedDict([('id', u'6'), ('pgrp', u'invscout'), ('groups', u'invscout'), ('home', u'/var/adm/invscout'), ('shell', u'/usr/bin/ksh')])), ('snapp', OrderedDict([('id', u'200'), ('pgrp', u'snapp'), ('groups', u'snapp'), ('home', u'/usr/sbin/snapp'), ('shell', u'/usr/sbin/snappd'), ('gecos', u'snapp login user')])), ('ipsec', OrderedDict([('id', u'230'), ('pgrp', u'staff'), ('groups', u'staff'), ('home', u'/etc/ipsec'), ('shell', u'/usr/bin/ksh')])), ('nuucp', OrderedDict([('id', u'7'), ('pgrp', u'uucp'), ('groups', u'uucp'), ('home', u'/var/spool/uucppublic'), ('shell', u'/usr/sbin/uucp/uucico'), ('gecos', u'uucp login user')])), ('pconsole', OrderedDict([('id', u'8'), ('pgrp', u'system'), ('groups', u'system,pconsole'), ('home', u'/var/adm/pconsole'), ('shell', u'/usr/bin/ksh')])), ('srvproxy', OrderedDict([('id', u'202'), ('pgrp', u'system'), ('groups', u'system,staff'), ('home', u'/home/srvproxy'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Service Proxy Daemon')])), ('esaadmin', OrderedDict([('id', u'10'), ('pgrp', u'system'), ('groups', u'system,staff'), ('home', u'/var/esa'), ('shell', u'/usr/bin/ksh')])), ('sshd', OrderedDict([('id', u'203'), ('pgrp', u'sshd'), ('groups', u'sshd,staff'), ('home', u'/var/empty'), ('shell', u'/usr/bin/ksh')])), ('nagios', OrderedDict([('id', u'5000'), ('pgrp', u'nagios'), ('groups', u'nagios,staff'), ('home', u'/usr/local/nagios'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Nagios Admin User')])), ('critter', OrderedDict([('id', u'888'), ('pgrp', u'knadmin'), ('groups', u'knadmin'), ('home', u'/local/home/critter'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Carsten Ritter, GI-ID')])), ('ujanssen', OrderedDict([('id', u'330'), ('pgrp', u'knadmin'), ('groups', u'knadmin'), ('home', u'/local/home/ujanssen'), ('shell', u'/usr/bin/bash'), ('gecos', u'Urs Janssen, GI-ID')])), ('nketelse', OrderedDict([('id', u'355'), ('pgrp', u'knadmin'), ('groups', u'knadmin'), ('home', u'/local/home/nketelse'), ('shell', u'/usr/bin/bash'), ('gecos', u'Nils Ketelsen, GI-ID')])), ('ohannula', OrderedDict([('id', u'11032'), ('pgrp', u'knadmin'), ('groups', u'knadmin'), ('home', u'/local/home/ohannula'), ('shell', u'/usr/bin/bash'), ('gecos', u'Olev Hannula, GI-ID')])), ('rwolter', OrderedDict([('id', u'1000001'), ('pgrp', u'knadmin'), ('groups', u'knadmin'), ('home', u'/local/home/rwolter'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Rudolf Wolter, GI-ID')])), ('kharrak', OrderedDict([('id', u'8396000'), ('pgrp', u'knadmin'), ('groups', u'knadmin'), ('home', u'/local/home/kharrak'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Karl-Hans Arrak, GI-ID')])), ('hholst', OrderedDict([('id', u'83956021'), ('pgrp', u'knadmin'), ('groups', u'knadmin'), ('home', u'/local/home/hholst'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Henrik Holst, HAM GI-ID')])), ('edi', OrderedDict([('id', u'201'), ('pgrp', u'edi'), ('groups', u'edi,ediops'), ('home', u'/home/edi'), ('shell', u'/usr/bin/ksh'), ('gecos', u'edi')])), ('ansiblec', OrderedDict([('id', u'501'), ('pgrp', u'ansible'), ('groups', u'ansible,staff'), ('home', u'/home/ansiblec'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Ansible Client User'), ('roles', u'')]))]))])), ('dehamsp20', OrderedDict([('exec_rc', 'rc=0'), ('exec_msg', 'SUCCESS'), ('users', OrderedDict([('root', OrderedDict([('id', u'0'), ('pgrp', u'system'), ('groups', u'system,bin,sys,security,cron,audit,lp'), ('home', u'/'), ('shell', u'/usr/bin/ksh')])), ('daemon', OrderedDict([('id', u'1'), ('pgrp', u'staff'), ('groups', u'staff,daemon'), ('home', u'/etc')])), ('bin', OrderedDict([('id', u'2'), ('pgrp', u'bin'), ('groups', u'bin,sys,adm'), ('home', u'/bin')])), ('sys', OrderedDict([('id', u'3'), ('pgrp', u'sys'), ('groups', u'sys'), ('home', u'/usr/sys')])), ('adm', OrderedDict([('id', u'4'), ('pgrp', u'adm'), ('groups', u'adm'), ('home', u'/var/adm')])), ('guest', OrderedDict([('id', u'100'), ('pgrp', u'usr'), ('groups', u'usr'), ('home', u'/home/guest')])), ('nobody', OrderedDict([('id', u'4294967294'), ('pgrp', u'nobody'), ('groups', u'nobody'), ('home', u'/')])), ('lpd', OrderedDict([('id', u'9'), ('pgrp', u'nobody'), ('groups', u'nobody'), ('home', u'/')])), ('lp', OrderedDict([('id', u'11'), ('pgrp', u'aconn'), ('groups', u'aconn,printq,lp'), ('home', u'/var/spool/lp'), ('shell', u'/bin/false')])), ('invscout', OrderedDict([('id', u'10'), ('pgrp', u'mqm'), ('groups', u'mqm,invscout'), ('home', u'/var/adm/invscout'), ('shell', u'/usr/bin/ksh')])), ('ipsec', OrderedDict([('id', u'314'), ('pgrp', u'staff'), ('groups', u'staff'), ('home', u'/etc/ipsec'), ('shell', u'/usr/bin/ksh')])), ('sshd', OrderedDict([('id', u'8'), ('pgrp', u'sshd'), ('groups', u'sshd,staff'), ('home', u'/var/empty'), ('shell', u'/usr/bin/ksh')])), ('edi', OrderedDict([('id', u'201'), ('pgrp', u'edi'), ('groups', u'edi,staff,fas,remote,mqm,ediinst'), ('home', u'/home/edi'), ('shell', u'/bin/ksh'), ('gecos', u'edi')])), ('edi0', OrderedDict([('id', u'301'), ('pgrp', u'edi'), ('groups', u'edi,staff,mqm,ediinst'), ('home', u'/home/edi0'), ('shell', u'/usr/bin/ksh'), ('gecos', u'edi0')])), ('edi2', OrderedDict([('id', u'303'), ('pgrp', u'edi'), ('groups', u'edi,staff,ediinst'), ('home', u'/home/edi2'), ('shell', u'/usr/bin/ksh'), ('gecos', u'edi2')])), ('edi3', OrderedDict([('id', u'321'), ('pgrp', u'edi'), ('groups', u'edi,staff,mqm,ediinst'), ('home', u'/home/edi3'), ('shell', u'/usr/bin/ksh'), ('gecos', u'edi3')])), ('edi4', OrderedDict([('id', u'323'), ('pgrp', u'edi'), ('groups', u'edi,staff,mqm,ediinst'), ('home', u'/home/edi4'), ('shell', u'/usr/bin/ksh'), ('gecos', u'edi4')])), ('edi5', OrderedDict([('id', u'380'), ('pgrp', u'edi'), ('groups', u'edi,staff,mqm,ediinst'), ('home', u'/home/edi5'), ('shell', u'/usr/bin/ksh'), ('gecos', u'edi5')])), ('extdr', OrderedDict([('id', u'243'), ('pgrp', u'edi'), ('groups', u'edi,staff,gscpom,ediops'), ('home', u'/home/extdr'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Daniel Ritter')])), ('micsw', OrderedDict([('id', u'211'), ('pgrp', u'edi'), ('groups', u'edi,gscpom,ediops'), ('home', u'/home/micsw'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Sven Willenbuecher')])), ('miejf', OrderedDict([('id', u'298'), ('pgrp', u'edi'), ('groups', u'edi,staff,gscpom,ediops'), ('home', u'/home/miejf'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Jurij Fajnberg')])), ('mieref', OrderedDict([('id', u'209'), ('pgrp', u'edi'), ('groups', u'edi,staff'), ('home', u'/home/mieref'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Referenz Account')])), ('oracle', OrderedDict([('id', u'332'), ('pgrp', u'dba'), ('groups', u'dba,edi'), ('home', u'/home/oracle'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Oracle admin')])), ('uucp', OrderedDict([('id', u'5'), ('pgrp', u'uucp'), ('groups', u'uucp,staff'), ('home', u'/usr/lib/uucp'), ('shell', u'/usr/bin/ksh'), ('gecos', u'DE;S;;213375;;Hermann Handick;;OS')])), ('nagios', OrderedDict([('id', u'5000'), ('pgrp', u'nagios'), ('groups', u'nagios,mqm'), ('home', u'/usr/local/nagios'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Nagios Admin User')])), ('critter', OrderedDict([('id', u'888'), ('pgrp', u'knadmin'), ('groups', u'knadmin,staff'), ('home', u'/home/critter'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Carsten Ritter, GI-ID')])), ('edi1', OrderedDict([('id', u'226'), ('pgrp', u'edi'), ('groups', u'edi,staff,mqm,ediinst'), ('home', u'/home/edi1'), ('shell', u'/usr/bin/ksh'), ('gecos', u'edi1')])), ('ujanssen', OrderedDict([('id', u'330'), ('pgrp', u'knadmin'), ('groups', u'knadmin,system,staff,security,shutdown'), ('home', u'/home/ujanssen'), ('shell', u'/usr/bin/bash'), ('gecos', u'Urs Janssen, GI-ID')])), ('nketelse', OrderedDict([('id', u'355'), ('pgrp', u'knadmin'), ('groups', u'knadmin,system,staff,security,shutdown'), ('home', u'/home/nketelse'), ('shell', u'/usr/bin/bash'), ('gecos', u'Nils Ketelsen, GI-ID')])), ('ohannula', OrderedDict([('id', u'11032'), ('pgrp', u'knadmin'), ('groups', u'knadmin,system,staff,security,shutdown'), ('home', u'/home/ohannula'), ('shell', u'/usr/bin/bash'), ('gecos', u'Olev Hannula, GI-ID')])), ('edi6', OrderedDict([('id', u'408'), ('pgrp', u'edi'), ('groups', u'edi,staff,mqm,ediinst'), ('home', u'/home/edi6'), ('shell', u'/usr/bin/ksh'), ('gecos', u'edi6')])), ('extrr', OrderedDict([('id', u'409'), ('pgrp', u'edi'), ('groups', u'edi,staff,gscpom,ediops'), ('home', u'/home/extrr'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Reinhold Redel')])), ('extto', OrderedDict([('id', u'410'), ('pgrp', u'edi'), ('groups', u'edi,staff,gscpom,ediops'), ('home', u'/home/extto'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Theo Ohnsorge')])), ('rwolter', OrderedDict([('id', u'1000001'), ('pgrp', u'knadmin'), ('groups', u'knadmin,system,staff,security,shutdown'), ('home', u'/home/rwolter'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Rudolf Wolter, GI-ID')])), ('miehy', OrderedDict([('id', u'433'), ('pgrp', u'edi'), ('groups', u'edi,gscpom,ediops'), ('home', u'/home/miehy'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Halberd Yao')])), ('kharrak', OrderedDict([('id', u'8396000'), ('pgrp', u'knadmin'), ('groups', u'knadmin,system,staff,security,shutdown'), ('home', u'/home/kharrak'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Karl-Hans Arrak, GI-ID')])), ('extjr', OrderedDict([('id', u'83956014'), ('pgrp', u'edi'), ('groups', u'edi,staff,gscpom'), ('home', u'/home/extjr'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Joachim Raue')])), ('hholst', OrderedDict([('id', u'461'), ('pgrp', u'knadmin'), ('groups', u'knadmin,staff'), ('home', u'/home/hholst'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Henrik Holst, HAM GI-ID')])), ('snapp', OrderedDict([('id', u'462'), ('pgrp', u'snapp'), ('groups', u'snapp'), ('home', u'/usr/sbin/snapp'), ('shell', u'/usr/bin/ksh'), ('gecos', u'snapp login user')])), ('extjg', OrderedDict([('id', u'464'), ('pgrp', u'edi'), ('groups', u'edi,gscpom,ediops'), ('home', u'/home/extjg'), ('shell', u'/usr/bin/ksh'), ('gecos', u'external.jacek.golek@kuehne-nagel.com')])), ('extms', OrderedDict([('id', u'465'), ('pgrp', u'edi'), ('groups', u'edi,staff'), ('home', u'/home/extms'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Michael Schoene')])), ('extbl', OrderedDict([('id', u'466'), ('pgrp', u'edi'), ('groups', u'edi,staff,gscpom,ediops'), ('home', u'/home/extbl'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Bernd Laser')])), ('gscpom', OrderedDict([('id', u'467'), ('pgrp', u'edi'), ('groups', u'edi,staff,fas,remote,mqm'), ('home', u'/home/gscpom'), ('shell', u'/usr/bin/ksh'), ('gecos', u'gscpom edi user')])), ('hkentgen', OrderedDict([('id', u'468'), ('pgrp', u'knadmin'), ('groups', u'knadmin,staff'), ('home', u'/home/hkentgen'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Heinz Kentgens, GI-ID')])), ('extrn', OrderedDict([('id', u'83956030'), ('pgrp', u'edi'), ('groups', u'edi,staff,gscpom,ediops'), ('home', u'/home/extrn'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Ren\xe9 Neumann')])), ('ansiblec', OrderedDict([('id', u'501'), ('pgrp', u'ansible'), ('groups', u'ansible,staff'), ('home', u'/home/ansiblec'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Ansible Client User for OSY'), ('roles', u'')])), ('extfs', OrderedDict([('id', u'64605'), ('pgrp', u'edi'), ('groups', u'edi,staff,gscpom,ediops'), ('home', u'/home/extfs'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Frank Seitz')])), ('extsi', OrderedDict([('id', u'64606'), ('pgrp', u'staff'), ('groups', u'staff'), ('home', u'/home/extsi'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Stanislav Ignatev')])), ('extas', OrderedDict([('id', u'64607'), ('pgrp', u'staff'), ('groups', u'staff'), ('home', u'/home/extas'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Alexander Slobodenyuk')])), ('extati', OrderedDict([('id', u'64608'), ('pgrp', u'staff'), ('groups', u'staff'), ('home', u'/home/extati'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Aleksei Tirman')])), ('extjs', OrderedDict([('id', u'64609'), ('pgrp', u'staff'), ('groups', u'staff'), ('home', u'/home/extjs'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Jishu Sharma')])), ('tllak', OrderedDict([('id', u'83956039'), ('pgrp', u'edi'), ('groups', u'edi'), ('home', u'/home/tllak'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Alo Kobin')])), ('ftpedi', OrderedDict([('id', u'475'), ('pgrp', u'edi'), ('groups', u'edi,staff'), ('home', u'/ib/dat/download/ftp'), ('shell', u'/usr/bin/ksh')])), ('jenkins', OrderedDict([('id', u'10254'), ('pgrp', u'jenkins'), ('groups', u'jenkins,staff'), ('home', u'/home/jenkins'), ('shell', u'/usr/bin/ksh')])), ('tllaa', OrderedDict([('id', u'249'), ('pgrp', u'edi'), ('groups', u'edi'), ('home', u'/home/tllaa'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Aleksandr Arepjev')])), ('tllsp', OrderedDict([('id', u'407'), ('pgrp', u'edi'), ('groups', u'edi,cron,gscpom,ediops'), ('home', u'/home/tllsp'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Sven Pajumagi')])), ('tlltl', OrderedDict([('id', u'229'), ('pgrp', u'edi'), ('groups', u'edi,cron,gscpom,ediops'), ('home', u'/home/tlltl'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Tanel Likk')])), ('tllpl', OrderedDict([('id', u'447'), ('pgrp', u'edi'), ('groups', u'edi,cron,gscpom'), ('home', u'/home/tllpl'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Priit Lohmus')])), ('tlljm', OrderedDict([('id', u'15183'), ('pgrp', u'edi'), ('groups', u'edi,cron,gscpom'), ('home', u'/home/tlljm'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Jakov Motsenov')])), ('nginx', OrderedDict([('id', u'64600'), ('pgrp', u'nginx'), ('groups', u'nginx,staff,edi'), ('home', u'/home/nginx'), ('shell', u'/usr/bin/ksh'), ('gecos', u'NGINX User')])), ('extrs', OrderedDict([('id', u'64613'), ('pgrp', u'edi'), ('groups', u'edi,staff,gscpom,ediops'), ('home', u'/home/extrs'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Ronald Sedelies')])), ('miecg', OrderedDict([('id', u'64610'), ('pgrp', u'edi'), ('groups', u'edi,staff,gscpom,ediops'), ('home', u'/home/miecg'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Christian Gebauer')])), ('extnl', OrderedDict([('id', u'64615'), ('pgrp', u'edi'), ('groups', u'edi,gscpom,ediops'), ('home', u'/home/extnl'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Neven Luetic')]))]))])), ('dehamsp36', OrderedDict([('exec_rc', 'rc=0'), ('exec_msg', 'SUCCESS'), ('users', OrderedDict([('root', OrderedDict([('id', u'0'), ('pgrp', u'system'), ('groups', u'system,bin,sys,security,cron,audit,lp'), ('home', u'/'), ('shell', u'/usr/bin/ksh')])), ('daemon', OrderedDict([('id', u'1'), ('pgrp', u'staff'), ('groups', u'staff'), ('home', u'/etc')])), ('bin', OrderedDict([('id', u'2'), ('pgrp', u'bin'), ('groups', u'bin,sys,adm'), ('home', u'/bin')])), ('sys', OrderedDict([('id', u'3'), ('pgrp', u'sys'), ('groups', u'sys'), ('home', u'/usr/sys')])), ('adm', OrderedDict([('id', u'4'), ('pgrp', u'adm'), ('groups', u'adm'), ('home', u'/var/adm')])), ('uucp', OrderedDict([('id', u'5'), ('pgrp', u'uucp'), ('groups', u'uucp'), ('home', u'/usr/lib/uucp')])), ('guest', OrderedDict([('id', u'100'), ('pgrp', u'usr'), ('groups', u'usr'), ('home', u'/home/guest')])), ('nobody', OrderedDict([('id', u'4294967294'), ('pgrp', u'nobody'), ('groups', u'nobody'), ('home', u'/')])), ('lpd', OrderedDict([('id', u'9'), ('pgrp', u'nobody'), ('groups', u'nobody'), ('home', u'/')])), ('lp', OrderedDict([('id', u'11'), ('pgrp', u'lp'), ('groups', u'lp,printq'), ('home', u'/var/spool/lp'), ('shell', u'/bin/false')])), ('invscout', OrderedDict([('id', u'6'), ('pgrp', u'invscout'), ('groups', u'invscout'), ('home', u'/var/adm/invscout'), ('shell', u'/usr/bin/ksh')])), ('snapp', OrderedDict([('id', u'200'), ('pgrp', u'snapp'), ('groups', u'snapp'), ('home', u'/usr/sbin/snapp'), ('shell', u'/usr/sbin/snappd'), ('gecos', u'snapp login user')])), ('ipsec', OrderedDict([('id', u'201'), ('pgrp', u'staff'), ('groups', u'staff'), ('home', u'/etc/ipsec'), ('shell', u'/usr/bin/ksh')])), ('nuucp', OrderedDict([('id', u'7'), ('pgrp', u'uucp'), ('groups', u'uucp'), ('home', u'/var/spool/uucppublic'), ('shell', u'/usr/sbin/uucp/uucico'), ('gecos', u'uucp login user')])), ('pconsole', OrderedDict([('id', u'8'), ('pgrp', u'system'), ('groups', u'system,pconsole'), ('home', u'/var/adm/pconsole'), ('shell', u'/usr/bin/ksh')])), ('esaadmin', OrderedDict([('id', u'10'), ('pgrp', u'system'), ('groups', u'system,staff'), ('home', u'/var/esa'), ('shell', u'/usr/bin/ksh')])), ('sshd', OrderedDict([('id', u'202'), ('pgrp', u'sshd'), ('groups', u'sshd,staff'), ('home', u'/var/empty'), ('shell', u'/usr/bin/ksh')])), ('nagios', OrderedDict([('id', u'5000'), ('pgrp', u'nagios'), ('groups', u'nagios'), ('home', u'/usr/local/nagios'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Nagios Admin user')])), ('hmcscan', OrderedDict([('id', u'9999'), ('pgrp', u'staff'), ('groups', u'staff'), ('home', u'/opt/hmcscanner'), ('shell', u'/usr/bin/ksh'), ('gecos', u'User for HMCscanner tool')])), ('rwolter', OrderedDict([('id', u'1000001'), ('pgrp', u'knadmin'), ('groups', u'knadmin,staff'), ('home', u'/home/rwolter'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Rudolf Wolter, GD-ID')])), ('critter', OrderedDict([('id', u'203'), ('pgrp', u'knadmin'), ('groups', u'knadmin,staff'), ('home', u'/home/critter'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Carsten Ritter, GI-ID')])), ('emrichp', OrderedDict([('id', u'204'), ('pgrp', u'knadmin'), ('groups', u'knadmin'), ('home', u'/home/emrichp'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Peter Emrich')])), ('srvproxy', OrderedDict([('id', u'205'), ('pgrp', u'system'), ('groups', u'system,staff'), ('home', u'/home/srvproxy'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Service Proxy Daemon')])), ('hholst', OrderedDict([('id', u'206'), ('pgrp', u'knadmin'), ('groups', u'knadmin,staff'), ('home', u'/home/hholst'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Henrik Holst, HAM GI-ID')])), ('hkentgen', OrderedDict([('id', u'207'), ('pgrp', u'knadmin'), ('groups', u'knadmin,staff'), ('home', u'/home/hkentgen'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Heinz Kentgens, GI-ID')])), ('ansiblec', OrderedDict([('id', u'501'), ('pgrp', u'ansible'), ('groups', u'ansible,staff'), ('home', u'/home/ansiblec'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Ansible Client User for OSY'), ('roles', u'')]))]))])), ('denotsp36', OrderedDict([('exec_rc', 'rc=0'), ('exec_msg', 'SUCCESS'), ('users', OrderedDict([('root', OrderedDict([('id', u'0'), ('pgrp', u'system'), ('groups', u'system,bin,sys,security,cron,audit,lp'), ('home', u'/'), ('shell', u'/usr/bin/ksh')])), ('daemon', OrderedDict([('id', u'1'), ('pgrp', u'staff'), ('groups', u'staff'), ('home', u'/etc')])), ('bin', OrderedDict([('id', u'2'), ('pgrp', u'bin'), ('groups', u'bin,sys,adm'), ('home', u'/bin')])), ('sys', OrderedDict([('id', u'3'), ('pgrp', u'sys'), ('groups', u'sys'), ('home', u'/usr/sys')])), ('adm', OrderedDict([('id', u'4'), ('pgrp', u'adm'), ('groups', u'adm'), ('home', u'/var/adm')])), ('uucp', OrderedDict([('id', u'5'), ('pgrp', u'uucp'), ('groups', u'uucp'), ('home', u'/usr/lib/uucp')])), ('guest', OrderedDict([('id', u'100'), ('pgrp', u'usr'), ('groups', u'usr'), ('home', u'/home/guest')])), ('nobody', OrderedDict([('id', u'4294967294'), ('pgrp', u'nobody'), ('groups', u'nobody'), ('home', u'/')])), ('lpd', OrderedDict([('id', u'9'), ('pgrp', u'nobody'), ('groups', u'nobody'), ('home', u'/')])), ('lp', OrderedDict([('id', u'11'), ('pgrp', u'lp'), ('groups', u'lp,printq'), ('home', u'/var/spool/lp'), ('shell', u'/bin/false')])), ('invscout', OrderedDict([('id', u'6'), ('pgrp', u'invscout'), ('groups', u'invscout'), ('home', u'/var/adm/invscout'), ('shell', u'/usr/bin/ksh')])), ('snapp', OrderedDict([('id', u'200'), ('pgrp', u'snapp'), ('groups', u'snapp'), ('home', u'/usr/sbin/snapp'), ('shell', u'/usr/sbin/snappd'), ('gecos', u'snapp login user')])), ('ipsec', OrderedDict([('id', u'201'), ('pgrp', u'staff'), ('groups', u'staff'), ('home', u'/etc/ipsec'), ('shell', u'/usr/bin/ksh')])), ('nuucp', OrderedDict([('id', u'7'), ('pgrp', u'uucp'), ('groups', u'uucp'), ('home', u'/var/spool/uucppublic'), ('shell', u'/usr/sbin/uucp/uucico'), ('gecos', u'uucp login user')])), ('pconsole', OrderedDict([('id', u'8'), ('pgrp', u'system'), ('groups', u'system,pconsole'), ('home', u'/var/adm/pconsole'), ('shell', u'/usr/bin/ksh')])), ('esaadmin', OrderedDict([('id', u'10'), ('pgrp', u'system'), ('groups', u'system,staff'), ('home', u'/var/esa'), ('shell', u'/usr/bin/ksh')])), ('sshd', OrderedDict([('id', u'202'), ('pgrp', u'sshd'), ('groups', u'sshd,staff'), ('home', u'/var/empty'), ('shell', u'/usr/bin/ksh')])), ('nagios', OrderedDict([('id', u'10101'), ('pgrp', u'nagios'), ('groups', u'nagios'), ('home', u'/usr/local/nagios'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Nagios Admin user')])), ('rwolter', OrderedDict([('id', u'1000001'), ('pgrp', u'knadmin'), ('groups', u'knadmin,staff'), ('home', u'/home/rwolter'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Rudolf Wolter, GD-ID')])), ('critter', OrderedDict([('id', u'203'), ('pgrp', u'knadmin'), ('groups', u'knadmin,staff'), ('home', u'/home/critter'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Carsten Ritter, GI-ID')])), ('emrichp', OrderedDict([('id', u'204'), ('pgrp', u'knadmin'), ('groups', u'knadmin,staff'), ('home', u'/home/emrichp'), ('shell', u'/usr/bin/ksh')])), ('srvproxy', OrderedDict([('id', u'205'), ('pgrp', u'system'), ('groups', u'system,staff'), ('home', u'/home/srvproxy'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Service Proxy Daemon')])), ('hholst', OrderedDict([('id', u'206'), ('pgrp', u'knadmin'), ('groups', u'knadmin,staff'), ('home', u'/home/hholst'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Henrik Holst, HAM GI-ID')])), ('ansiblec', OrderedDict([('id', u'501'), ('pgrp', u'ansible'), ('groups', u'ansible,staff'), ('home', u'/home/ansiblec'), ('shell', u'/usr/bin/ksh'), ('gecos', u'Ansible Client User for OSY'), ('roles', u'')]))]))]))])
	users = order_by_user(raw_users)

	# Preparing the HTML Static content
	html_file = open('../tools/blueprints/page/templates/users.html', 'w+')
	html_file.write("{% extends 'layouts/base.html' %}\n")
	html_file.write('{% block title %} Easy	Manager - Users	{% endblock %}\n')
	html_file.write('\n')
	html_file.write('{% block body %}\n')
	excel_file = read_excel('em_lsusers.xlsx', None)
	#users = excel_file.keys() # list containing all users retrieved by Ansible

	### GENERATING THE HTML CODE ###
	html_file.write('<div class="container" >\n')

	# Creating the Report Button
	html_file.write('   <a href="#Report" class ="btn btn-primary" role="button">Report</a>\n')

	# Creating the dropdown Button with user filter
	html_file.write('   <div class="dropdown">\n')
	html_file.write('       <button class="btn btn-primary dropdown-toggle" type="button" data-toggle="dropdown">Users\n')
	html_file.write('       <span class="caret"></span></button>\n')
	html_file.write('       <ul class="dropdown-menu">\n')
	html_file.write('           <input class="form-control" id="userFilter" type="text" placeholder="Filter...">\n')
	for user in users.iterkeys():
		html_file.write('           <li ><a data-toggle="pill" href="#{}">{}</a></li>\n'.format(user, user))
	html_file.write('       </ul>\n')
	html_file.write('   </div>\n')

	# looping over each user's tab and creating the appropriate HTML code
	html_file.write('   <div class="tab-content">\n')
	for user in users.iterkeys():
		html_file.write('       <div id="{}" class="tab-pane">\n'.format(user))
		#html = read_excel('em_lsusers.xlsx', user).to_html().encode('utf-8')
		#print html
		html_file.write(mk_html_table(users[user])) # Parsing the table's HTML code.
		html_file.write('\n')
		html_file.write('       </div>\n')
	html_file.write('   </div>\n')
	html_file.write('</div>\n')
	html_file.write('<script>\n')
	html_file.write('   $(document).ready(function(){\n')
	html_file.write('       $("#userFilter").on("keyup", function() {\n')
	html_file.write('           var value = $(this).val().toLowerCase();\n')
	html_file.write('           $(".dropdown-menu li").filter(function() {\n')
	html_file.write('               $(this).toggle($(this).text().toLowerCase().indexOf(value) > -1)\n')
	html_file.write('           });\n')
	html_file.write('       });\n')
	html_file.write('   });\n')
	html_file.write('</script>\n')
	### END OF THE HTML CODE ###

	html_file.write("{% endblock %}\n")
	html_file.close()


# --------------------------------------------------------------- #
### END OF FUNCTIONS DECLARATION